import openpyxl

wb_obj = openpyxl.load_workbook("OKO.xlsx")
sheet_obj = wb_obj.active

"""
cell_obj = sheet_obj.cell(row = 3, column = 2)
kriteri = sheet_obj.cell(row = 3, column = 1)

print(cell_obj.value) - названия столбцов
print(sheet_obj.max_row) - количество строк
print(sheet_obj.max_column) - количество столбцов
"""
#печать первого значения столбца
m = sheet_obj.max_row
m = m+1 
i = 3
while i < m: 
    pokazatel = sheet_obj.cell(row = i, column = 3)
    kriteri = sheet_obj.cell(row = i, column = 1)
    print(str(kriteri.value) + ' - ' + str(pokazatel.value))
    i = i+1